"""
Agent 30 — Vision Intelligence Agent
OCR and document intelligence using GPT-4 Vision / Claude Vision for Bengali documents.
Handles scanned PDFs, handwritten BOQ, and Bengali text extraction.
"""

from __future__ import annotations

import base64
import json
import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

from .base import BaseAgent, AgentResult, AgentStatus

logger = logging.getLogger(__name__)


class VisionIntelligenceAgent(BaseAgent):
    agent_id = "agent-029-vision-intelligence"
    agent_name = "Vision Intelligence Agent"
    description = "Extracts text from scanned Bengali tender documents using GPT-4 Vision / Claude Vision OCR. Handles handwritten BOQ, faded print, and complex tables."
    dependencies: List[str] = ["agent-004-document-ai"]
    version = "1.0.0"

    def __init__(self):
        super().__init__()
        self.openai_key = os.getenv("OPENAI_API_KEY", "")
        self.anthropic_key = os.getenv("ANTHROPIC_API_KEY", "")

    async def execute(self, context: Dict[str, Any]) -> AgentResult:
        file_path = context.get("file_path", "")
        document_type = context.get("document_type", "boq")  # boq, notice, tds, general
        language = context.get("language", "bn")  # bn, en, mix
        use_vision = context.get("use_vision", True)
        
        if not file_path or not Path(file_path).exists():
            return AgentResult(
                agent_id=self.agent_id,
                agent_name=self.agent_name,
                status=AgentStatus.FAILED,
                output={"error": f"File not found: {file_path}"},
            )

        # Extract text using vision or fallback methods
        extracted = await self._extract_document(file_path, document_type, language, use_vision)

        return AgentResult(
            agent_id=self.agent_id,
            agent_name=self.agent_name,
            status=AgentStatus.SUCCESS,
            output=extracted,
        )

    async def _extract_document(
        self, file_path: str, doc_type: str, language: str, use_vision: bool
    ) -> Dict[str, Any]:
        """Extract document content using best available method."""
        ext = Path(file_path).suffix.lower()
        result = {"file_path": file_path, "document_type": doc_type, "method": "fallback"}
        
        # Try Vision API first (if key available)
        if use_vision and self.openai_key:
            try:
                vision_result = await self._extract_with_vision_api(file_path, language)
                if vision_result.get("success"):
                    result.update(vision_result)
                    result["method"] = "gpt4_vision"
                    return result
            except Exception as e:
                logger.warning(f"Vision API failed: {e}")

        # Try Claude Vision as fallback
        if use_vision and self.anthropic_key:
            try:
                claude_result = await self._extract_with_claude_vision(file_path, language)
                if claude_result.get("success"):
                    result.update(claude_result)
                    result["method"] = "claude_vision"
                    return result
            except Exception as e:
                logger.warning(f"Claude Vision failed: {e}")

        # Local OCR fallback
        fallback = await self._extract_local(file_path, language)
        result.update(fallback)
        result["method"] = "local_ocr"
        return result

    async def _extract_with_vision_api(self, file_path: str, language: str) -> Dict:
        """Extract text using OpenAI GPT-4 Vision API."""
        try:
            from openai import OpenAI
            client = OpenAI(api_key=self.openai_key)
            
            # Read and encode image
            with open(file_path, "rb") as f:
                base64_image = base64.b64encode(f.read()).decode("utf-8")
            
            ext = Path(file_path).suffix.lower()
            media_type = "image/png" if ext == ".png" else "image/jpeg" if ext in (".jpg", ".jpeg") else "application/pdf"
            
            language_instruction = (
                "The document is in Bengali (Bangla). Extract all text exactly as written."
                if language == "bn" else
                "The document contains both Bengali and English text. Extract all text preserving the original language."
                if language == "mix" else
                "The document is in English. Extract all text exactly as written."
            )
            
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are a precise OCR engine for Bangladeshi government tender documents. "
                            "Extract ALL text from the document exactly as written, preserving:\n"
                            "- Numbers and quantities (identify as numeric values)\n"
                            "- Table structures (output as structured JSON for tables)\n"
                            "- Bengali text (preserve Unicode)\n"
                            "- Currency amounts in BDT\n"
                            "- Dates, tender IDs, and reference numbers\n\n"
                            "Output JSON format:\n"
                            "{\n"
                            '  "full_text": "extracted text",\n'
                            '  "tables": [{headers, rows}],\n'
                            '  "key_values": {"field": "value"},\n'
                            '  "confidence": 0.95\n'
                            "}"
                        ),
                    },
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": f"Extract all text from this {doc_type} document. {language_instruction}"},
                            {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{base64_image}"}},
                        ],
                    },
                ],
                max_tokens=4096,
                response_format={"type": "json_object"},
            )
            
            result = json.loads(response.choices[0].message.content)
            result["success"] = True
            result["model"] = "gpt-4o"
            result["tokens_used"] = response.usage.total_tokens
            
            return result
            
        except ImportError:
            logger.warning("openai package not installed")
            return {"success": False, "error": "openai package not installed"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    async def _extract_with_claude_vision(self, file_path: str, language: str) -> Dict:
        """Extract text using Claude Vision API."""
        try:
            import anthropic
            
            client = anthropic.Anthropic(api_key=self.anthropic_key)
            
            with open(file_path, "rb") as f:
                image_data = base64.b64encode(f.read()).decode("utf-8")
            
            ext = Path(file_path).suffix.lower()
            media_type = "image/png" if ext == ".png" else "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
            
            msg = client.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=4096,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": (
                                    f"Extract all text from this Bangladeshi tender {doc_type} document. "
                                    "The document may contain Bengali text. Extract everything exactly. "
                                    "Output as JSON with: full_text, tables (as structured data), key_values (field:value), confidence"
                                ),
                            },
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": media_type,
                                    "data": image_data,
                                },
                            },
                        ],
                    }
                ],
            )
            
            content = msg.content[0].text if msg.content else "{}"
            # Try to parse as JSON
            try:
                result = json.loads(content)
            except Exception:
                result = {"full_text": content, "tables": [], "key_values": {}, "confidence": 0.7}
            
            result["success"] = True
            result["model"] = "claude-3.5-sonnet"
            
            return result
            
        except ImportError:
            logger.warning("anthropic package not installed")
            return {"success": False, "error": "anthropic package not installed"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    async def _extract_local(self, file_path: str, language: str) -> Dict:
        """Extract text using local OCR (Tesseract + pdfplumber fallback)."""
        ext = Path(file_path).suffix.lower()
        full_text = ""
        tables = []
        key_values = {}
        
        try:
            if ext == ".pdf":
                # Try pdfplumber for digital PDFs
                try:
                    import pdfplumber
                    with pdfplumber.open(file_path) as pdf:
                        for page in pdf.pages:
                            text = page.extract_text() or ""
                            full_text += text + "\n"
                            
                            # Try table extraction
                            page_tables = page.extract_tables()
                            for t in page_tables:
                                if t:
                                    tables.append({
                                        "headers": t[0] if t else [],
                                        "rows": t[1:] if len(t) > 1 else [],
                                    })
                except Exception:
                    pass

                # Fallback to PyPDF2
                if not full_text.strip():
                    try:
                        from PyPDF2 import PdfReader
                        reader = PdfReader(file_path)
                        for page in reader.pages:
                            full_text += (page.extract_text() or "") + "\n"
                    except Exception:
                        pass

            elif ext in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
                # Try Tesseract OCR for images
                try:
                    import pytesseract
                    from PIL import Image
                    img = Image.open(file_path)
                    lang = "ben+eng" if language in ("bn", "mix") else "eng"
                    full_text = pytesseract.image_to_string(img, lang=lang)
                except ImportError:
                    full_text = "[OCR not available - install pytesseract and tesseract-ocr-ben]"
                except Exception as e:
                    full_text = f"[OCR failed: {e}]"
            
            elif ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        sheet_rows = []
                        for row in ws.iter_rows(values_only=True):
                            sheet_rows.append([str(c) if c is not None else "" for c in row])
                        if sheet_rows:
                            tables.append({
                                "sheet_name": sheet_name,
                                "headers": sheet_rows[0],
                                "rows": sheet_rows[1:],
                            })
                except Exception:
                    pass

        except Exception as e:
            logger.error(f"Local extraction failed: {e}")
        
        # Extract key-value pairs from text
        if full_text:
            import re
            patterns = [
                (r"Tender\s*[Ii][Dd]\s*[:\-]?\s*(\S+)", "tender_id"),
                (r"(?:BDT|টাকা|Tk\.?)\s*([\d,]+\.?\d*)", "amount_bdt"),
                (r"(\d{2}[/\-]\d{2}[/\-]\d{4})", "date"),
                (r"ই-জিপি\s*[:\-]?\s*(\S+)", "egp_id"),
            ]
            for pattern, key in patterns:
                match = re.search(pattern, full_text)
                if match:
                    key_values[key] = match.group(1)
        
        return {
            "success": bool(full_text.strip()),
            "full_text": full_text.strip() or "[No text extracted]",
            "tables": tables,
            "key_values": key_values,
            "confidence": 0.6 if full_text.strip() else 0.0,
        }
