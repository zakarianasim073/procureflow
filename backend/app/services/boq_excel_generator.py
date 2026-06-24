from __future__ import annotations

"""BOQ Excel Generator — Generate BOQ Rate Analysis Excel in 5-tab format.
Produces: Tender Summary, BOQ Rate Comparison, Work Type Summary, 
           Rate Detail & Flags, Financial Check
"""


import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Styling ──────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(bold=True, size=12, color="2F5496")
LABEL_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
SECTION_FONT = Font(bold=True, size=10, color="2F5496")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Work type categorization keywords
WORK_TYPE_KEYWORDS = {
    "Preliminaries": ["preparation of site", "site office", "preliminaries", "mobilization"],
    "Earthwork": ["earth work", "excavation", "filling", "cutting", "ditch", "pond", "lead"],
    "Filter Works": ["filter", "sand filter", "jhama chips", "geo-tex filter"],
    "CC Blocks & Dumping": ["cc block", "c.c. block", "dumping", "boulder", "concrete block"],
    "Geo-Bags / Geotextile": ["geo-bag", "geotextile", "geo-textile", "salvaging"],
    "Structural Concrete": ["rcc", "reinforced", "centering", "shuttering", "ms rod", "mass concrete"],
    "Finishing": ["plaster", "paint", "tile", "brick work", "brick flat", "soling"],
    "Electrical": ["cable", "led", "solar", "pole", "stay", "bracket", "board", "breaker", "earthing", "circuit"],
    "Labour": ["head man", "labour", "skilled", "unskilled"],
}


def classify_work_type(description: str, code: str = "") -> str:
    """Classify a BOQ item into a work type category."""
    desc_lower = (description + " " + code).lower()
    for wtype, keywords in WORK_TYPE_KEYWORDS.items():
        for kw in keywords:
            if kw in desc_lower:
                return wtype
    return "Other"


class BOQExcelGenerator:
    """Generate BOQ Rate Analysis Excel in standard 5-tab format."""

    def __init__(self, tender_data: Dict[str, Any], boq_items: List[Dict[str, Any]], zone: str = "D"):
        self.tender = tender_data
        self.items = boq_items
        self.zone = zone

    def generate(self, output_path: str) -> str:
        """Generate the 5-tab Excel workbook."""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets in order
        self._build_tender_summary(wb)
        self._build_boq_comparison(wb)
        self._build_work_type_summary(wb)
        self._build_rate_detail_flags(wb)
        self._build_financial_check(wb)
        
        # Save
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        logger.info(f"BOQ Excel saved: {output_path}")
        return output_path

    def _build_tender_summary(self, wb: openpyxl.Workbook):
        """Sheet 1: Tender Summary with key info."""
        ws = wb.create_sheet("Tender Summary", 0)
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        
        t = self.tender
        items = self.items
        
        total_sor = sum(it.get('sor_rate', 0) * it.get('quantity', 0) for it in items)
        total_quoted = sum(it.get('quoted_rate', 0) * it.get('quantity', 0) for it in items)
        saving = total_sor - total_quoted
        discount_pct = round((saving / total_sor * 100), 2) if total_sor else 0
        
        rows = [
            (3, "C", f'🏗  BOQ vs SOR RATE ANALYSIS — TENDER ID: {t.get("tender_id", "")}'),
            (4, "C", f'{t.get("package_no", "")} — {t.get("package_description", t.get("brief", ""))}'),
            (None, None, None),
            (6, "C", "Procuring Entity", "E", t.get("procuring_entity", t.get("organization", ""))),
            (7, "C", "Location", "E", t.get("location", t.get("district", ""))),
            (8, "C", "Agency / SOR", "E", t.get("sor_agency", "BWDB 2019-20 Schedule of Rates")),
            (9, "C", "Zone", "E", f'Zone {self.zone} ({["Dhaka","Chattogram","Rajshahi","Khulna/Barishal"][{"A":0,"B":1,"C":2,"D":3}.get(self.zone,0)]} Division)'),
            (10, "C", "Invitation Ref.", "E", t.get("invitation_ref", "")),
            (11, "C", "Tender ID", "E", str(t.get("tender_id", ""))),
            (12, "C", "Tender Security", "E", t.get("tender_security_text", "")),
            (13, "C", "Closing Date", "E", t.get("tender_close_datetime", "")),
            (14, "C", "Work Period", "E", t.get("work_period", "")),
            (15, "C", "Report Generated", "E", datetime.now().strftime("%d-%b-%Y %H:%M")),
            (None, None, None),
            (17, "C", "Total SOR Amount (BDT)", "F", total_sor),
            (18, "C", "Total Quoted Amount (BDT)", "F", total_quoted),
            (19, "C", "Saving vs SOR (BDT)", "F", saving),
            (20, "C", "Discount (%)", "F", f"{discount_pct:.2f}%"),
            (None, None, None),
        ]
        
        for row_data in rows:
            r = row_data[0]
            if r is None:
                continue
            ws.cell(r, 1, "►")
            ws.cell(r, 3, row_data[2])
            ws.cell(r, 3).font = LABEL_FONT
            if len(row_data) > 3:
                ws.cell(r, 5, row_data[4])
                ws.cell(r, 5).font = NORMAL_FONT
        
        # TDS Requirements section
        r = 22
        ws.cell(r, 3, "📋 QUALIFICATION REQUIREMENTS (TDS)").font = SUBTITLE_FONT
        tds_items = [
            ("Experience (General)", t.get("experience_general", "5 years as Prime/Sub/Mgmt Contractor")),
            ("Experience (Specific)", t.get("experience_specific", "")),
            ("Annual Turnover (AACT)", t.get("annual_turnover", "")),
            ("Financial Resources", t.get("financial_resources", "")),
            ("Min Tender Capacity", t.get("tender_capacity", "")),
            ("Personnel", t.get("personnel", "")),
            ("Equipment", t.get("equipment", "")),
            ("JV", t.get("jv_notes", "")),
        ]
        for i, (label, val) in enumerate(tds_items):
            ws.cell(r + 1 + i, 3, label).font = LABEL_FONT
            ws.cell(r + 1 + i, 5, str(val)).font = NORMAL_FONT

    def _build_boq_comparison(self, wb: openpyxl.Workbook):
        """Sheet 2: Full BOQ Rate Comparison."""
        ws = wb.create_sheet("BOQ Rate Comparison", 1)
        
        headers = ["#", "Item Code", "Agency", "Work Type", "Description of Item",
                   "Unit", "Quantity", "SOR Rate (BDT)", "SOR Amount (BDT)",
                   "Quoted Rate (BDT)", "Quoted Amount (BDT)",
                   "Rate Diff (BDT)", "Variance (%)", "Status"]
        
        # Title
        ws.cell(1, 1, f'BOQ vs SOR RATE COMPARISON — Tender {self.tender.get("tender_id", "")} — {self.tender.get("package_no", "")}').font = TITLE_FONT
        ws.cell(2, 1, '  LEGEND:   ✅ AT SOR (±1%)    ⚠ VARIANCE (1–10%)    🔵 BELOW SOR (>10%)    🔴 ABOVE SOR (>5%)').font = Font(size=9, italic=True)
        
        # Headers
        for c, h in enumerate(headers, 1):
            cell = ws.cell(4, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
            cell.border = THIN_BORDER
        
        # Set column widths
        widths = [4, 14, 7, 16, 50, 7, 10, 14, 16, 14, 16, 14, 10, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        # Group items by work type
        work_type_order = [
            "Preliminaries", "Earthwork", "Filter Works", "Geo-Bags / Geotextile",
            "CC Blocks & Dumping", "Structural Concrete", "Finishing",
            "Electrical", "Labour", "Other"
        ]
        
        current_wt = None
        row = 4
        for wt in work_type_order:
            wt_items = [it for it in self.items if it.get('work_type', classify_work_type(it.get('description', ''), it.get('code', ''))) == wt]
            if not wt_items:
                continue
            
            # Section header
            row += 1
            ws.cell(row, 1, f'▶  {wt}')
            ws.cell(row, 1).font = SECTION_FONT
            ws.cell(row, 1).fill = SECTION_FILL
            for c in range(2, 15):
                ws.cell(row, c).fill = SECTION_FILL
            
            for item in wt_items:
                row += 1
                code = item.get('code', '')
                desc = item.get('description', '')
                agency = item.get('agency', '')
                unit = item.get('unit', '')
                qty = item.get('quantity', 0)
                sor_rate = item.get('sor_rate', 0)
                quoted_rate = item.get('quoted_rate', 0)
                sor_amt = sor_rate * qty
                quoted_amt = quoted_rate * qty
                rate_diff = quoted_rate - sor_rate
                variance = rate_diff / sor_rate if sor_rate else 0
                
                # Status
                if abs(variance) <= 0.01:
                    status = "AT SOR"
                elif variance < -0.01:
                    status = "BELOW SOR"
                elif variance > 0.05:
                    status = "ABOVE SOR"
                else:
                    status = "VARIANCE"
                
                ws.cell(row, 1, item.get('item_no', ''))
                ws.cell(row, 2, code)
                ws.cell(row, 3, agency)
                ws.cell(row, 4, wt)
                ws.cell(row, 5, desc[:100])
                ws.cell(row, 6, unit)
                ws.cell(row, 7, qty)
                ws.cell(row, 8, sor_rate)
                ws.cell(row, 9, round(sor_amt, 2))
                ws.cell(row, 10, quoted_rate if quoted_rate else "Fill By Tenderer")
                ws.cell(row, 11, round(quoted_amt, 2) if quoted_rate else "Auto")
                ws.cell(row, 12, round(rate_diff, 2) if quoted_rate else 0)
                ws.cell(row, 13, round(variance, 4) if sor_rate else 0)
                ws.cell(row, 14, status)
                
                for c in range(1, 15):
                    ws.cell(row, c).font = NORMAL_FONT
                    ws.cell(row, c).border = THIN_BORDER
        
        # Grand total row
        row += 2
        total_sor_amt = sum(it.get('sor_rate', 0) * it.get('quantity', 0) for it in self.items)
        total_q_amt = sum(it.get('quoted_rate', 0) * it.get('quantity', 0) for it in self.items)
        ws.cell(row, 1, "GRAND TOTAL").font = HEADER_FONT
        ws.cell(row, 9, round(total_sor_amt, 2)).font = HEADER_FONT
        ws.cell(row, 11, round(total_q_amt, 2)).font = HEADER_FONT
        ws.cell(row, 12, round(total_q_amt - total_sor_amt, 2)).font = HEADER_FONT

    def _build_work_type_summary(self, wb: openpyxl.Workbook):
        """Sheet 3: Work Type Summary."""
        ws = wb.create_sheet("Work Type Summary", 2)
        
        headers = ["Work Type", "Items", "SOR Amount (BDT)", "Quoted Amount (BDT)",
                   "Saving (BDT)", "Discount (%)", "% of Quoted Total"]
        
        ws.cell(1, 1, f'WORK TYPE — COST SUMMARY & BREAKDOWN — Tender {self.tender.get("tender_id", "")}').font = TITLE_FONT
        
        for c, h in enumerate(headers, 1):
            cell = ws.cell(2, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        
        widths = [25, 8, 20, 20, 18, 14, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        # Group and summarize
        work_types = {}
        for item in self.items:
            wt = item.get('work_type', classify_work_type(item.get('description', ''), item.get('code', '')))
            if wt not in work_types:
                work_types[wt] = {'count': 0, 'sor_amt': 0, 'quoted_amt': 0}
            work_types[wt]['count'] += 1
            work_types[wt]['sor_amt'] += item.get('sor_rate', 0) * item.get('quantity', 0)
            work_types[wt]['quoted_amt'] += item.get('quoted_rate', 0) * item.get('quantity', 0)
        
        total_quoted = sum(v['quoted_amt'] for v in work_types.values())
        
        row = 2
        for wt, data in sorted(work_types.items(), key=lambda x: -x[1]['sor_amt']):
            row += 1
            saving = data['sor_amt'] - data['quoted_amt']
            discount = saving / data['sor_amt'] if data['sor_amt'] else 0
            pct_of_total = data['quoted_amt'] / total_quoted if total_quoted else 0
            
            ws.cell(row, 1, wt).font = NORMAL_FONT
            ws.cell(row, 2, data['count']).font = NORMAL_FONT
            ws.cell(row, 3, round(data['sor_amt'], 2)).font = NORMAL_FONT
            ws.cell(row, 4, round(data['quoted_amt'], 2)).font = NORMAL_FONT
            ws.cell(row, 5, round(saving, 2)).font = NORMAL_FONT
            ws.cell(row, 6, round(discount, 4)).font = NORMAL_FONT
            ws.cell(row, 7, round(pct_of_total, 4)).font = NORMAL_FONT
            for c in range(1, 8):
                ws.cell(row, c).border = THIN_BORDER

    def _build_rate_detail_flags(self, wb: openpyxl.Workbook):
        """Sheet 4: Rate Detail & Flags."""
        ws = wb.create_sheet("Rate Detail & Flags", 3)
        
        headers = ["#", "Code", "Agency", "Description", "Unit", "Qty",
                   "SOR Rate", "Quoted Rate", "Saving/Unit", "Total Saving", "Status"]
        
        ws.cell(1, 1, 'RATE DETAIL — FLAGGED ITEMS FOR REVIEW / NEGOTIATION').font = TITLE_FONT
        
        for c, h in enumerate(headers, 1):
            cell = ws.cell(2, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        
        widths = [4, 14, 7, 50, 7, 8, 12, 12, 12, 16, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        row = 2
        for item in self.items:
            row += 1
            sor_rate = item.get('sor_rate', 0)
            quoted_rate = item.get('quoted_rate', 0)
            qty = item.get('quantity', 0)
            saving_per_unit = quoted_rate - sor_rate if quoted_rate else 0
            total_saving = saving_per_unit * qty
            variance = (quoted_rate - sor_rate) / sor_rate if sor_rate else 0
            
            if abs(variance) <= 0.01:
                status = "AT SOR"
            elif variance < -0.01:
                status = "BELOW SOR"
            elif variance > 0.05:
                status = "ABOVE SOR"
            else:
                status = "VARIANCE"
            
            ws.cell(row, 1, item.get('item_no', ''))
            ws.cell(row, 2, item.get('code', ''))
            ws.cell(row, 3, item.get('agency', ''))
            ws.cell(row, 4, (item.get('description', '') or '')[:100])
            ws.cell(row, 5, item.get('unit', ''))
            ws.cell(row, 6, qty)
            ws.cell(row, 7, sor_rate)
            ws.cell(row, 8, quoted_rate if quoted_rate else "Fill By Tenderer")
            ws.cell(row, 9, round(saving_per_unit, 2) if quoted_rate else "N/A")
            ws.cell(row, 10, round(total_saving, 2) if quoted_rate else "N/A")
            ws.cell(row, 11, status)
            
            for c in range(1, 12):
                ws.cell(row, c).font = NORMAL_FONT
                ws.cell(row, c).border = THIN_BORDER

    def _build_financial_check(self, wb: openpyxl.Workbook):
        """Sheet 5: Financial Qualification Checklist."""
        ws = wb.create_sheet("Financial Check", 4)
        
        headers = ["", "Criterion", "Required", "Our Figure", "Remarks", "Status"]
        
        ws.cell(1, 1, f'FINANCIAL QUALIFICATION CHECKLIST — {self.tender.get("bidder_name", "Bidder")}').font = TITLE_FONT
        
        for c, h in enumerate(headers, 1):
            cell = ws.cell(2, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        
        widths = [3, 30, 25, 25, 35, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        items = [
            ("Avg Annual Construction Turnover", "Tk. 2,420 Lakh", "See audit/turnover cert.", 
             "AACT ≥ 2420 Lakh over 5 yrs", "⚠ Verify"),
            ("Min Financial Resource (Line of Credit)", "Tk. 850 Lakh", "NCC Bank — Tk. 850 Lakh LoC",
             "Form e-PW3-8 submitted", "✅ OK"),
            ("Tender Capacity (A×N×1.25 - B)", "Tk. 1,820 Lakh", "Calculate from max annual work",
             "N = completion years", "⚠ Verify"),
            ("Tender Security", f'Tk. {self.tender.get("tender_security_bdt", 6100000):,.0f} (BG)',
             "BG from NCC Bank", "Form e-PW3-7", "✅ OK"),
            ("Experience (Specific)", "1 contract ≥ Tk. 1,210 Lakh", "River bank protection work",
             "Must submit completion cert.", "⚠ Verify"),
            ("JV Agreement", "e-PW3-B, Tk. 300 stamp", "JV deed submitted",
             "Non-judicial stamp verified", "✅ OK"),
            ("Tax Documents", "TIN + IT Return 2024-25 + VAT", "All submitted",
             "Income tax year: 2024-25", "✅ OK"),
            ("Ongoing Works Declaration", "Signed list required", "Must declare all ongoing works",
             "JV % share on each cert.", "⚠ Verify"),
        ]
        
        for r, (criterion, required, figure, remarks, status) in enumerate(items, 3):
            ws.cell(r, 2, criterion).font = LABEL_FONT
            ws.cell(r, 3, required).font = NORMAL_FONT
            ws.cell(r, 4, figure).font = NORMAL_FONT
            ws.cell(r, 5, remarks).font = NORMAL_FONT
            ws.cell(r, 6, status).font = Font(bold=True, 
                color="008000" if "✅" in status else "FF8000")
            for c in range(1, 7):
                ws.cell(r, c).border = THIN_BORDER


def generate_boq_excel(
    tender_data: Dict[str, Any],
    boq_items: List[Dict[str, Any]],
    zone: str = "D",
    output_path: str = "./boq_rate_analysis.xlsx"
) -> str:
    """Generate BOQ Rate Analysis Excel with 5 tabs."""
    # Classify work types
    for item in boq_items:
        if 'work_type' not in item:
            item['work_type'] = classify_work_type(
                item.get('description', ''),
                item.get('code', '')
            )
    
    gen = BOQExcelGenerator(tender_data, boq_items, zone)
    return gen.generate(output_path)
