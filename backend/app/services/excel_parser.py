"""
Excel Parser Service
Extracts BOQ items and SOR rates from Excel files
"""

import re
from typing import List, Dict, Optional
from pathlib import Path
import openpyxl
from app.core import helpers
from app.core.match_helpers import normalize_code, normalize_text


class ExcelParser:
    """Parse BOQ and SOR data from Excel files"""

    async def extract_boq_items(self, excel_path: str) -> List[Dict]:
        """Extract BOQ line items from Excel"""
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        items = []
        headers = {}
        in_table = False

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            vals = [c.value for c in row]

            if any(h in str(vals).lower() for h in ['item', 'code', 'description', 'item no']):
                for i, v in enumerate(vals):
                    if v:
                        v_lower = str(v).lower().strip()
                        if 'item' in v_lower and ('no' in v_lower or '#' in v_lower or 'number' in v_lower):
                            headers['item_no'] = i
                        elif 'code' in v_lower:
                            headers['code'] = i
                        elif 'desc' in v_lower:
                            headers['desc'] = i
                        elif 'unit' in v_lower:
                            headers['unit'] = i
                        elif 'qty' in v_lower or 'quant' in v_lower:
                            headers['qty'] = i
                        elif 'rate' in v_lower:
                            headers['rate'] = i
                in_table = bool(headers)
                continue

            if not in_table:
                continue

            item_no = str(vals[headers.get('item_no', 0)] or '').strip() if headers.get('item_no') is not None else ''
            code = str(vals[headers.get('code', 0)] or '').strip() if headers.get('code') is not None else ''
            desc = str(vals[headers.get('desc', 0)] or '').strip() if headers.get('desc') is not None else ''
            unit = str(vals[headers.get('unit', 0)] or '').strip() if headers.get('unit') is not None else ''
            qty = helpers.to_num(vals[headers.get('qty', 0)]) if headers.get('qty') is not None else None
            rate = helpers.to_num(vals[headers.get('rate', 0)]) if headers.get('rate') is not None else None

            if not code and not desc:
                continue
            if code and not code.replace('-', '').replace('.', '').isdigit():
                continue

            items.append({
                "item_no": item_no,
                "code": code,
                "description": desc,
                "unit": unit.lower(),
                "quantity": qty,
                "rate": rate
            })

        wb.close()
        return items

    async def extract_sor_rates(
        self,
        excel_path: str,
        zone: Optional[str] = None,
        agency: str = "BWDB"
    ) -> Dict[str, float]:
        """Extract SOR rates from Excel with optional zone selection"""
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        rates = {}
        headers = {}
        in_table = False

        zone_keys = {
            'A': 'zone_a', 'B': 'zone_b', 'C': 'zone_c', 'D': 'zone_d',
            'a': 'zone_a', 'b': 'zone_b', 'c': 'zone_c', 'd': 'zone_d'
        }

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            vals = [c.value for c in row]

            if any(h in str(vals).lower() for h in ['code', 'item code', 'sor code', 'rate']):
                for i, v in enumerate(vals):
                    if v:
                        v_lower = str(v).lower().strip()
                        if 'code' in v_lower:
                            headers['code'] = i
                        elif 'desc' in v_lower:
                            headers['desc'] = i
                        elif 'unit' in v_lower:
                            headers['unit'] = i
                        elif 'rate' in v_lower and 'zone' not in v_lower:
                            headers['rate'] = i
                        elif v_lower in ('a', 'zone a', 'zone-a'):
                            headers['zone_a'] = i
                        elif v_lower in ('b', 'zone b', 'zone-b'):
                            headers['zone_b'] = i
                        elif v_lower in ('c', 'zone c', 'zone-c'):
                            headers['zone_c'] = i
                        elif v_lower in ('d', 'zone d', 'zone-d'):
                            headers['zone_d'] = i
                in_table = True
                continue

            if not in_table:
                continue

            code = str(vals[headers.get('code', 0)] or '').strip() if headers.get('code') is not None else ''
            desc = str(vals[headers.get('desc', 0)] or '').strip() if headers.get('desc') is not None else ''

            if not code and not desc:
                continue

            rate = None
            zone_col = zone_keys.get(zone) if zone else None
            if zone_col and zone_col in headers:
                rate = helpers.to_num(vals[headers[zone_col]])
            elif headers.get('rate') is not None:
                rate = helpers.to_num(vals[headers['rate']])

            if rate and code:
                rates[code.strip()] = float(rate)

        wb.close()
        return rates
