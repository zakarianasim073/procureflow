"""Full end-to-end test: zone dict + APP estimate + TDS criteria"""
import httpx, json

resp = httpx.post(
    "http://localhost:8000/api/boq/compare",
    data={
        "boq_file_id": "57c20c63",
        "sor_agency": "BWDB",
        "zone": '{"BWDB":"B","PWD":"B","LGED":"B"}',
        "tender_info": json.dumps({
            "tender_id": "1290886",
            "title": "Ban-61/2025-26, Construction of Regulator (3-Vent: 1.50m x 1.80m) on Kumuria Khal at Sonali Bazar",
            "entity": "Cox's Bazar WD Division 2",
            "location": "Pekua, Coxsbazar",
        }),
    },
    timeout=60,
)
print(f"Status: {resp.status_code}")
if resp.status_code == 200:
    d = resp.json()
    print(f"Success: {d.get('success')}")
    print(f"Items: {d.get('total_items')}")
    print(f"Excel: {d.get('excel_path')}")
    print(f"Summary total_sor: {d['summary'].get('total_sor')}")
    print(f"Summary total_quoted: {d['summary'].get('total_quoted')}")
    print(f"Keys: {list(d.keys())}")

    # Check the Excel file
    import openpyxl
    xl_path = d['excel_path']
    if not xl_path.startswith("D:\\"):
        xl_path = f"D:\\A1\\procurementflow_final_v3\\procurementflow\\backend\\{xl_path}"
    wb = openpyxl.load_workbook(xl_path)
    print(f"\nSheets: {wb.sheetnames}")
    ws1 = wb['Tender Summary']
    print("\nTender Summary:")
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, values_only=True):
        vals = [str(v)[:60] if v else "" for v in row]
        print(f"  {vals}")

    # Check Sheet 2 for Remarks column
    ws2 = wb['BOQ Rate Comparison']
    print("\nBOQ Rate Comparison (first 10 data rows):")
    for row in ws2.iter_rows(min_row=3, max_row=min(13, ws2.max_row), values_only=True):
        vals = [str(v)[:40] if v else "" for v in row]
        print(f"  {vals}")

    # Check SOR NOT FOUND items
    print("\nSOR NOT FOUND items:")
    found = 0
    for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, values_only=True):
        if row[-1] == 'SOR NOT FOUND':
            print(f"  Code={row[1]}, Desc={str(row[4])[:50]}")
            found += 1
    if found == 0:
        print("  (none)")
    print(f"\nTotal SOR NOT FOUND: {found}")

    ws5 = wb['Financial Check']
    print("\nFinancial Check:")
    for row in ws5.iter_rows(min_row=1, max_row=ws5.max_row, values_only=True):
        vals = [str(v)[:40] if v else "" for v in row]
        print(f"  {vals}")

    # Check flagged items
    ws4 = wb['Rate Detail & Flags']
    print("\nFlagged items:")
    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, values_only=True):
        if any(v for v in row):
            vals = [str(v)[:40] if v else "" for v in row]
            print(f"  {vals}")
else:
    print(resp.text[:500])
